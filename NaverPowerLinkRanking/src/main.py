# pyinstaller
# pyinstaller --clean --onefile --noconsole --icon=icon/icon.ico main.py

'''
[.spec]
import sys
sys.setrecursionlimit(1500)

[compt.py]
out = out.decode(encoding, errors='ignore')
'''

# pyinstaller --clean --onefile --noconsole --icon=icon/icon.ico main.spec

import sys
import os
import json
import re
import csv
import chardet
import time
import threading
import requests
from time import sleep

import ErrorPopup_Tkinter as ePopup
import FileOpen_easygui as fopen
import URL_KeywordParser as urlParser
import MakeCSV

from tkinter import *
from tkinter.ttk import *
import tkinter.messagebox

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import QtCore, QtGui, QtWidgets

from urllib.parse import quote_plus
from openpyxl import load_workbook as load_workbook

# '기본' 탭에서 사용됨.
url_col = []        # 테이블에 보여줄 URL주소를 저장
keyword_col = []    # 테이블에 보여줄 키워드를 저장
ranking_col = []    # 테이블에 보여줄 순위를 저장
searchBtnCnt = 0    # 검색 버튼이 몇회 눌렸는지 카운팅
PageCnt = 1         # '기본' 탭에서 현재 페이지 수를 표시

# '대량조회' 탭에서 사용됨.
url_col2 = []       # 테이블에 보여줄 URL주소를 저장
keyword_col2 = []   # 테이블에 보여줄 키워드를 저장
ranking_col2 = []   # 테이블에 보여줄 순위를 저장
lineCnt = 0         # CSV Data Line Counter
multiPageCnt = 1    # '대량조회' 탭에서 현재 페이지 수를 표시
totalLines = 0
multiSearchFilePath = ''

global net_checker  # 네트워크 상태를 확인

# Qt Designer 로 작성
class Ui_Dialog(object):    
    def setupUi(self, Dialog):
        Dialog.setObjectName("PowerLink Ranking Program")
        Dialog.resize(692, 892)
        self.tabWidget = QtWidgets.QTabWidget(Dialog)
        self.tabWidget.setGeometry(QtCore.QRect(10, 10, 671, 861))
        self.tabWidget.setObjectName("tabWidget")

        # 현재 네트워크 연결상태를 나타내는 Label
        self.netStateTitleLabel = QtWidgets.QLabel(Dialog)
        self.netStateTitleLabel.setGeometry(QtCore.QRect(400, 5, 90, 30))
        self.netStateTitleLabel.setObjectName("network_state_title")

        self.netStateLabel = QtWidgets.QLabel(Dialog)
        self.netStateLabel.setGeometry(QtCore.QRect(500, 5, 150, 30))
        self.netStateLabel.setObjectName("network_state")
        self.netStateLabel.setStyleSheet('color: #6E6E6E')

        ############################## Basic Tab ##############################
        self.Basic_Tab = QtWidgets.QWidget()
        self.Basic_Tab.setObjectName("Basic_Tab")

        self.label_Keyword = QtWidgets.QLabel(self.Basic_Tab)
        self.label_Keyword.setGeometry(QtCore.QRect(50, 70, 70, 20))
        self.label_Keyword.setObjectName("label_Keyword")

        self.label_URL = QtWidgets.QLabel(self.Basic_Tab)
        self.label_URL.setGeometry(QtCore.QRect(60, 30, 71, 21))
        self.label_URL.setObjectName("label_URL")

        self.Input_URL = QtWidgets.QLineEdit(self.Basic_Tab)
        self.Input_URL.setGeometry(QtCore.QRect(150, 20, 491, 31))
        self.Input_URL.setObjectName("Input_URL")
        self.Input_URL.setText("검색하려는 URL을 입력해주세요.")

        self.Input_Keyword = QtWidgets.QLineEdit(self.Basic_Tab)
        self.Input_Keyword.setGeometry(QtCore.QRect(150, 70, 491, 31))
        self.Input_Keyword.setObjectName("Input_Keyword")
        self.Input_Keyword.setText("해당 URL이 조회되기 위한 키워드를 입력해주세요.")

        self.Search = QtWidgets.QPushButton(self.Basic_Tab)
        self.Search.setGeometry(QtCore.QRect(520, 120, 121, 23))
        self.Search.setObjectName("Search")
        self.Search.clicked.connect(self.SearchBtnClicked)

        self.ResultTable = QtWidgets.QTableWidget(self.Basic_Tab)
        self.ResultTable.setGeometry(QtCore.QRect(20, 190, 621, 600))
        self.ResultTable.setObjectName("ResultTable")
        self.ResultTable.setColumnCount(3)      # 검색 결과를 보여줄 테이블의 열의 수를 지정
        self.ResultTable.setRowCount(25)        # 검색 결과를 보여줄 테이블의 행의 수를 지정
        column_headers = ['URL 주소', '키워드', '순위']     # 테이블 열의 이름을 지정
        self.ResultTable.setHorizontalHeaderLabels(column_headers)
        self.ResultTable.resizeColumnsToContents()
        self.ResultTable.resizeRowsToContents()

        self.Download = QtWidgets.QPushButton(self.Basic_Tab)
        self.Download.setGeometry(QtCore.QRect(520, 150, 121, 23))
        self.Download.setObjectName("Download")
        self.Download.clicked.connect(self.DownloadBtnClicked)

        self.Next = QtWidgets.QPushButton(self.Basic_Tab)
        self.Next.setGeometry(QtCore.QRect(420, 800, 20, 20))
        self.Next.setObjectName("Next")
        self.Next.clicked.connect(self.NextBtnClicked)

        self.pageLabel = QtWidgets.QLabel(self.Basic_Tab)
        self.pageLabel.setGeometry(QtCore.QRect(295, 800, 75, 23))
        self.pageLabel.setObjectName("pageLabel1")

        self.Previous = QtWidgets.QPushButton(self.Basic_Tab)
        self.Previous.setGeometry(QtCore.QRect(200, 800, 20, 20))
        self.Previous.setObjectName("Previous")
        self.Previous.clicked.connect(self.PreviousBtnClicked)

        self.tabWidget.addTab(self.Basic_Tab, "")

        self.resetTable = QtWidgets.QPushButton(self.Basic_Tab)
        self.resetTable.setGeometry(QtCore.QRect(390, 150, 121, 23))
        self.resetTable.setObjectName("Reset")
        self.resetTable.clicked.connect(self.ResetBtnClicked)

        ############################## Multi Tab ############################## 
        self.Multi_Tab = QtWidgets.QWidget()
        self.Multi_Tab.setObjectName("Multi_Tab")

        self.UploadLabel = QtWidgets.QLabel(self.Multi_Tab)
        self.UploadLabel.setGeometry(QtCore.QRect(40, 50, 75, 23))
        self.UploadLabel.setObjectName("Upload")

        self.LocalPath = QtWidgets.QLineEdit(self.Multi_Tab)
        self.LocalPath.setGeometry(QtCore.QRect(130, 40, 501, 31))
        self.LocalPath.setObjectName("LocalPath")
        self.LocalPath.setText(".xlsx 파일의 절대주소를 입력해주세요.")

        self.FileOpen = QtWidgets.QPushButton(self.Multi_Tab)
        self.FileOpen.setGeometry(QtCore.QRect(520, 90, 121, 23))
        self.FileOpen.setObjectName("FileOpen")
        self.FileOpen.clicked.connect(self.FileOpenBtnClicked)

        self.Search2 = QtWidgets.QPushButton(self.Multi_Tab)
        self.Search2.setGeometry(QtCore.QRect(520, 120, 121, 23))
        self.Search2.setObjectName("Search2")
        self.Search2.clicked.connect(self.Search2BtnClicked)

        self.Download2 = QtWidgets.QPushButton(self.Multi_Tab)
        self.Download2.setGeometry(QtCore.QRect(520, 150, 121, 23))
        self.Download2.setObjectName("Download2")
        self.Download2.clicked.connect(self.Download2BtnClicked)

        self.Previous2 = QtWidgets.QPushButton(self.Multi_Tab)
        self.Previous2.setGeometry(QtCore.QRect(200, 800, 20, 20))
        self.Previous2.setObjectName("Previous2")
        self.Previous2.clicked.connect(self.Previous2BtnClicked)

        self.pageLabel2 = QtWidgets.QLabel(self.Multi_Tab)
        self.pageLabel2.setGeometry(QtCore.QRect(295, 800, 75, 23))
        self.pageLabel2.setObjectName("pageLabel2")

        self.Next22 = QtWidgets.QPushButton(self.Multi_Tab)
        self.Next22.setGeometry(QtCore.QRect(420, 800, 20, 20))
        self.Next22.setObjectName("Next22")
        self.Next22.clicked.connect(self.Next2BtnClicked)

        self.ResultTable2 = QtWidgets.QTableWidget(self.Multi_Tab)
        self.ResultTable2.setGeometry(QtCore.QRect(20, 190, 620, 600))
        self.ResultTable2.setObjectName("ResultTable2")
        self.ResultTable2.setColumnCount(3)
        self.ResultTable2.setRowCount(25)
        self.ResultTable2.setHorizontalHeaderLabels(column_headers)
        self.ResultTable2.resizeColumnsToContents()
        self.ResultTable2.resizeRowsToContents()

        self.resetTable2 = QtWidgets.QPushButton(self.Multi_Tab)
        self.resetTable2.setGeometry(QtCore.QRect(390, 150, 121, 23))
        self.resetTable2.setObjectName("Reset2")
        self.resetTable2.clicked.connect(self.ResetBtn2Clicked)

        self.tabWidget.addTab(self.Multi_Tab, "")
        self.retranslateUi(Dialog)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

        app.aboutToQuit.connect(self.closeEvent)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate(
            "Dialog", "네이버 파워링크 순위 검색"))
        self.netStateTitleLabel.setText(_translate("Dialog","네트워크 상태 : "))
        self.netStateLabel.setText(_translate("Dialog","확인중..."))
        self.label_Keyword.setText(_translate("Dialog", "검색 키워드"))
        self.label_URL.setText(_translate("Dialog", "URL 주소"))
        self.Search.setText(_translate("Dialog", "검색"))
        self.Download.setText(_translate("Dialog", "다운로드 (.csv)"))
        self.Next.setText(_translate("Dialog", ">"))
        self.pageLabel.setText(_translate("Dialog", "1"))
        self.Previous.setText(_translate("Dialog", "<"))
        self.resetTable.setText(_translate("Dialog", "초기화"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(
            self.Basic_Tab), _translate("Dialog", "기본"))
        self.UploadLabel.setText(_translate("Dialog", "파일 경로"))
        self.Search2.setText(_translate("Dialog", "검색"))
        self.Download2.setText(_translate("Dialog", "다운로드 (.csv)"))
        self.Next22.setText(_translate("Dialog", ">"))
        self.pageLabel2.setText(_translate("Dialog", "1"))
        self.Previous2.setText(_translate("Dialog", "<"))
        self.resetTable2.setText(_translate("Dialog", "초기화"))
        self.FileOpen.setText(_translate("Dialog", "열기"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(
            self.Multi_Tab), _translate("Dialog", "대량조회"))

    # '기본' 탭의 '검색' 버튼이 눌렸을 때 반응하는 함수
    def SearchBtnClicked(self):
        global searchBtnCnt, PageCnt
        searchBtnCnt += 1
        #print(searchBtnCnt, PageCnt)    # Debuging Point

        # '검색' 버튼이 눌리는 횟수에 맞춰 결과 테이블의 페이징 처리
        PageCntStr = ''
        if (searchBtnCnt % 25 == 0):
            for i in range(1, searchBtnCnt//25+1):
                PageCntStr += str(i)
                PageCntStr += str(' ')
        else:
            for i in range(1, searchBtnCnt//25+2):
                PageCntStr += str(i)
                PageCntStr += str(' ')
        self.pageLabel.setText(PageCntStr)

        keyword = self.Input_Keyword.text()
        searchUrl = self.Input_URL.text()

        # URL 주소와 Keyword 를 매개변수로 크롤링하는 외부 함수 호출
        urlResult, keywordResult, rankingResult = urlParser.PowerLinkPaser(
            searchUrl, keyword)
        url_col.extend(urlResult)
        keyword_col.extend(keywordResult)
        ranking_col.extend(rankingResult)

        tableTempData = {
            'url_col': url_col,
            'keyword_col': keyword_col,
            'ranking_col': ranking_col
        }
        column_idx_lookup = {'url_col': 0, 'keyword_col': 1, 'ranking_col': 2}

        # 기존에 보여주던 검색 결과를 초기화 하고 새로운 검색결과까지 반영한 결과 테이블을 생성
        self.ResultTable.setRowCount(0)
        self.ResultTable.setRowCount(25)
        for k, v in tableTempData.items():
            col = column_idx_lookup[k]
            for row, val in enumerate(v):
                if(row >= (PageCnt-1)*25):
                    item = QTableWidgetItem(val)
                    self.ResultTable.setItem(row-(PageCnt-1)*25, col, item)
        self.ResultTable.resizeColumnsToContents()  # 텍스트의 길이에 맞게 재조정
        self.ResultTable.resizeRowsToContents()

    # '대량조회' 탭의 '열기' 버튼이 눌렸을 때 반응하는 함수
    def FileOpenBtnClicked(self):
        absPath = fopen.OpenWinFileExplorer()
        try:
            fileExtension = os.path.splitext(absPath)[1]
        except:
            ePopup.loadWrongPath(str(absPath))
        else:
            #print(fileExtension)   # 선택한 파일의 확장자명 확인
            if (fileExtension == '.xlsx'):
                print("Load Ok!")
            else:
                ePopup.FileLoadError()  # 선택한 파일의 확장자명이 다른 확장자일 경우 에러 출력
        finally:
            self.LocalPath.setText(absPath)

    # '대량조회' 탭의 '검색' 버튼이 눌렸을 때 반응하는 함수
    def Search2BtnClicked(self):
        keyword_col2.clear()
        ranking_col2.clear()
        url_col2.clear()
        self.ResultTable2.setRowCount(0)
        self.ResultTable2.setRowCount(25)
        self.ResultTable2.resizeColumnsToContents()
        self.ResultTable2.resizeRowsToContents()
        global multiSearchFilePath
        multiSearchFilePath = self.LocalPath.text()
        global lineCnt
            
        ProgressApp = tkApp()   # 진행도 현황을 보여주기 위한 화면 객체 생성
        ProgressApp.mainloop()

        PageCntStr = '1 '
        if((lineCnt//25) > 0):
            for i in range(2, lineCnt//25+2):
                PageCntStr += str(i)
                PageCntStr += str(' ')
        #print(PageCntStr)  # Debuging Point
        self.pageLabel2.setText(PageCntStr)

        tableTempData = {
            'url_col': url_col2,
            'keyword_col': keyword_col2,
            'ranking_col': ranking_col2
        }
        column_idx_lookup = {'url_col': 0,
                                'keyword_col': 1, 'ranking_col': 2}

        for k, v in tableTempData.items():
            col = column_idx_lookup[k]
            for row, val in enumerate(v):
                item = QTableWidgetItem(val)
                self.ResultTable2.setItem(row, col, item)
        self.ResultTable2.resizeColumnsToContents()
        self.ResultTable2.resizeRowsToContents()

    # '기본' 탭의 '다운로드' 버늩이 눌렸을 때 반응하는 함수
    def DownloadBtnClicked(self):
        tableTempData = {
            'URL': url_col,
            'KEYWORD': keyword_col,
            'RANKING': ranking_col
        }
        #print(tableTempData)   # Debuging Point

        current_path = os.path.dirname(os.path.realpath(__file__))
        download_path = current_path + r'\SearchResults'
        
        # 만들어질 디렉토리의 존재 여부 확인
        if not os.path.exists(download_path):
            os.makedirs(download_path)

        global searchBtnCnt, PageCnt
        if(searchBtnCnt == 0):
            # 테이블에 검색된 데이터가 없는 경우 에러 출력
            ePopup.NoneTableData()
        else:
            # '다운로드' 버튼이 눌리는 시점을 기준으로 날짜와 시간을 파일 이름에 반영
            output_file_name = download_path + '\SearchResults_' + \
                str(time.strftime('%y'+'-'+'%m'+'-'+'%d'+'_' +
                                  '%H'+'-'+'%M'+'-'+'%S')) + quote_plus(".csv")
            MakeCSV.MakeDownloadCSV(output_file_name, tableTempData)

            # 다운로드 파일로 결과 값들을 출력하고 난 뒤 남아 있는 데이터 초기화
            url_col.clear()
            keyword_col.clear()
            ranking_col.clear()
            self.ResultTable.setRowCount(0)
            searchBtnCnt = 0
            PageCnt = 1
            PageCntStr = ''
            for i in range(1, searchBtnCnt//25+2):
                PageCntStr += str(i)
                PageCntStr += str(' ')
            self.pageLabel.setText(PageCntStr)
            self.ResultTable.setRowCount(25)
            self.ResultTable.resizeColumnsToContents()
            self.ResultTable.resizeRowsToContents()

    # '대량조회' 탭의 '다운로드' 버튼이 눌렸을 때 반응하는 함수
    def Download2BtnClicked(self):
        tableTempData = {
            'URL': url_col2,
            'KEYWORD': keyword_col2,
            'RANKING': ranking_col2
        }
        #print(tableTempData)   # Debuging Point

        current_path = os.path.dirname(os.path.realpath(__file__))
        download_path = current_path + r'\MultiSearchResults'
        if not os.path.exists(download_path):
            os.makedirs(download_path)

        global lineCnt
        if(lineCnt == 0):
            ePopup.NoneTableData()
        else:
            output_file_name = download_path + '\MultiSearchResults_' + \
                str(time.strftime('%y'+'-'+'%m'+'-'+'%d'+'_' +
                                  '%H'+'-'+'%M'+'-'+'%S')) + quote_plus(".csv")
            MakeCSV.MakeDownloadCSV(output_file_name, tableTempData)
        
        # 다운로드 파일로 결과 값들을 출력하고 난 뒤 남아 있는 데이터 초기화
        url_col2.clear()
        keyword_col2.clear()
        ranking_col2.clear()
        self.ResultTable2.setRowCount(0)
        self.ResultTable2.setRowCount(25)
        self.ResultTable2.resizeColumnsToContents()
        self.ResultTable2.resizeRowsToContents()

    # '기본' 탭의 '>' 버튼이 눌렸을 때 반응하는 함수
    def NextBtnClicked(self):
        global PageCnt
        if(searchBtnCnt > (25*PageCnt)):
            self.ResultTable.setRowCount(0)
            self.ResultTable.setRowCount(25)
            tableTempData = {
                'url_col': url_col,
                'keyword_col': keyword_col,
                'ranking_col': ranking_col
            }
            column_idx_lookup = {'url_col': 0,
                                 'keyword_col': 1, 'ranking_col': 2}

            for k, v in tableTempData.items():
                col = column_idx_lookup[k]
                for row, val in enumerate(v):
                    if(row >= (25*PageCnt)):
                        item = QTableWidgetItem(val)
                        self.ResultTable.setItem(row-PageCnt*25, col, item)
            self.ResultTable.resizeColumnsToContents()
            self.ResultTable.resizeRowsToContents()
            PageCnt += 1    # 페이지의 카운터를 증가
        else:
            ePopup.PageBtnError()

    # '대량조회' 탭의 '>' 버튼이 눌렸을 때 반응하는 함수
    def Next2BtnClicked(self):
        #print("Next2BtnClicked")   # Debuging Point
        global multiPageCnt
        if(lineCnt > (25*multiPageCnt)):
            self.ResultTable2.setRowCount(0)
            self.ResultTable2.setRowCount(25)
            tableTempData = {
                'url_col': url_col2,
                'keyword_col': keyword_col2,
                'ranking_col': ranking_col2
            }
            column_idx_lookup = {'url_col': 0,
                                 'keyword_col': 1, 'ranking_col': 2}

            for k, v in tableTempData.items():
                col = column_idx_lookup[k]
                for row, val in enumerate(v):
                    if(row >= (25*multiPageCnt)):
                        item = QTableWidgetItem(val)
                        self.ResultTable2.setItem(
                            row-multiPageCnt*25, col, item)
            self.ResultTable2.resizeColumnsToContents()
            self.ResultTable2.resizeRowsToContents()
            multiPageCnt += 1
        else:
            ePopup.PageBtnError()

    # '기본' 탭의 '<' 버튼이 눌렸을 때 반응하는 함수
    def PreviousBtnClicked(self):
        global PageCnt
        if(PageCnt != 1):
            PageCnt -= 1    # 페이지의 카운터를 감소
            self.ResultTable.setRowCount(0)
            self.ResultTable.setRowCount(25)
            tableTempData = {
                'url_col': url_col,
                'keyword_col': keyword_col,
                'ranking_col': ranking_col
            }
            column_idx_lookup = {'url_col': 0,
                                 'keyword_col': 1, 'ranking_col': 2}

            for k, v in tableTempData.items():
                col = column_idx_lookup[k]
                for row, val in enumerate(v):
                    if(row >= (PageCnt-1)*25):
                        item = QTableWidgetItem(val)
                        self.ResultTable.setItem(row-(PageCnt-1)*25, col, item)
            self.ResultTable.resizeColumnsToContents()
            self.ResultTable.resizeRowsToContents()
        else:
            ePopup.PageBtnError()

    # '대량조회' 탭의 '<' 버튼이 눌렸을 때 반응하는 함수
    def Previous2BtnClicked(self):
        #print("Previous2BtnClicked")   # Debuging Point
        global multiPageCnt
        if(multiPageCnt != 1):
            multiPageCnt -= 1
            self.ResultTable2.setRowCount(0)
            self.ResultTable2.setRowCount(25)
            tableTempData = {
                'url_col': url_col2,
                'keyword_col': keyword_col2,
                'ranking_col': ranking_col2
            }
            column_idx_lookup = {'url_col': 0,
                                 'keyword_col': 1, 'ranking_col': 2}

            for k, v in tableTempData.items():
                col = column_idx_lookup[k]
                for row, val in enumerate(v):
                    if(row >= (multiPageCnt-1)*25):
                        item = QTableWidgetItem(val)
                        self.ResultTable2.setItem(
                            row-(multiPageCnt-1)*25, col, item)
            self.ResultTable2.resizeColumnsToContents()
            self.ResultTable2.resizeRowsToContents()
        else:
            ePopup.PageBtnError()

    # '기본' 탭의 '초기화' 버튼이 눌렸을 때 반응하는 함수
    def ResetBtnClicked(self):
        self.ResultTable.setRowCount(0)
        self.ResultTable.setRowCount(25)
        self.ResultTable.resizeColumnsToContents()
        self.ResultTable.resizeRowsToContents()
        global searchBtnCnt, PageCnt
        global url_col, keyword_col, ranking_col
        url_col.clear()
        keyword_col.clear()
        ranking_col.clear()
        searchBtnCnt = 0
        PageCnt = 1

    # '대양조회' 탭의 '초기화' 버튼이 눌렸을 때 반응하는 함수
    def ResetBtn2Clicked(self):
        self.ResultTable2.setRowCount(0)
        self.ResultTable2.setRowCount(25)
        self.ResultTable2.resizeColumnsToContents()
        self.ResultTable2.resizeRowsToContents()
        global multiPageCnt, lineCnt
        global url_col2, keyword_col2, ranking_col2
        url_col2.clear()
        keyword_col2.clear()
        ranking_col2.clear()
        lineCnt = 0
        multiPageCnt = 1

    # 1초마다 네트워크가 정상적으로 연결되어 있는지 확인
    def net_checker(self):
        global net_checker
        try:
            req = requests.head("https://ad.search.naver.com/search.naver?")
            print("Server: " + req.headers['server'])
            print("Content type: " + req.headers['content-type'])
            print("Date: " + req.headers['Date'])
        except:     # 네트워크가 연결되어 있지 않을 경우, 빨간색으로 문구 표시
            self.netStateLabel.setText('연결되어 있지 않습니다.')
            self.netStateLabel.setStyleSheet('color: #FF0000')
        else:       # 네트워크가 정상 연결되어 있을 경우, 녹색으로 문구 표시
            self.netStateLabel.setText('연결되었습니다.')
            self.netStateLabel.setStyleSheet('color: #31B404')
        finally:
            net_checker = threading.Timer(1, self.net_checker)  # 네트워크 연결 상태를 1초마다 확인한다.
            net_checker.start()

    # 프로그램이 닫히기 직전 수행해야 하는 작업들 지정
    def closeEvent(self):
        global net_checker
        print('EXIT')
        net_checker.cancel()    # 사용자가 닫기 버튼을 눌러 프로그램을 종료하기 직전에 네트워크 상태를 확인하는 스레드를 종료시킨다.
        sys.exit(0)


# Progress Check Thread
progress_var = 0

# '대량조회' 시 현재 진행되는 상황을 실시간으로 나타내줌
class tkApp(Tk):
    global totalLines

    def __init__(self):
        super().__init__()
        self.stopFlag = 0        # '대량조회' 검색 진행중 중지 탐지

        self.title("진행상황")
        self.geometry('{}x{}'.format(400, 150))
        self.cautionText = tkinter.StringVar()
        self.cautionText.set("\n검색 진행 중 강제로 종료하지 마세요.\n")
        self.cautionLabel = Label(self, textvariable=self.cautionText)
        self.cautionLabel.pack()
        self.pgText = tkinter.StringVar()
        self.pgText.set("\n진행 : ( 0/0 )")
        self.theLabel = Label(self, textvariable=self.pgText)
        self.theLabel.pack()
        self.timeText = tkinter.StringVar()
        self.timeText.set("진행시간 : ")
        self.timeLabel = Label(self, textvariable=self.timeText)
        self.timeLabel.pack()
        self.checkBtn = Button(
            self, text='확인', state='disabled', takefocus=False, command=self.destroy)
        self.stopBtn = Button(
            self, text='중지', takefocus=False, command=self.stop_command)
        self.stopBtn.pack()
        self.checkBtn.pack(side=RIGHT)
        self.progress = tkinter.ttk.Progressbar(
            self, variable=progress_var, mode="determinate")
        self.progress.pack(fill=X, expand=1)
        self.PgChanger()

    def stop_command(self):     # 중지 버튼이 눌려진 경우 동작하는 함수
        if self.stopFlag == 0:
            self.stopFlag=1

    def PgChanger(self):
        wrongFormError = 0
        startTime = time.time()     # 크롤링해오는 시간을 측정
        global lineCnt, multiSearchFilePath
        try:
            load_wb = load_workbook(multiSearchFilePath, data_only=True)
        except:
            print('cannot open : ', multiSearchFilePath)
            ePopup.SearchDisable()
        else:
            sheet = load_wb.worksheets[0]
            lineCnt = 0
            xlData = []
            for row in sheet.rows:
                xlData.append([row[0].value, row[1].value])
            for i, dt in enumerate(xlData):
                if(i == 0) & ((dt[0] != 'URL') | (dt[1] != 'KEYWORD')):
                    #print("Wrong Form!")
                    self.checkBtn['takefocus'] = TRUE
                    self.checkBtn['state'] = 'active'
                    self.update()
                    wrongFormError = 1
                    ePopup.loadWrongForm(multiSearchFilePath)
                    xlData.clear()
                    break
                else:
                    del xlData[0]
                    break
            totalLines = len(xlData)
            if(totalLines != 0):
                self.progress['maximum'] = totalLines-1

            global progress_var, stopFlag
            for i, dt in enumerate(xlData):
                lineCnt = i
                #print(i+1, dt[0], dt[1])    # check

                # URL 주소와 Keyword 를 매개변수로 크롤링하는 외부 함수 호출
                urlResult, keywordResult, rankingResult = urlParser.PowerLinkPaser(
                    dt[0], dt[1])
                url_col2.extend(urlResult)
                keyword_col2.extend(keywordResult)
                ranking_col2.extend(rankingResult)

                progress_var = lineCnt
                self.progress['value'] = progress_var
                progressText = '\n진행 : ('+str(lineCnt+1) + \
                    '/'+str(totalLines)+')'
                endTime = time.time() - startTime
                endTime = round(endTime, 2)
                tempTimeText = '진행시간 : ' + str(endTime)+' 초'
                self.timeText.set(tempTimeText)
                self.pgText.set(progressText)
                self.progress.update()
                if(self.stopFlag==1):   # 중지 버튼이 눌려진 경우 크롤링 중지
                    self.stopFlag=0
                    break;
            load_wb.close()
            if(wrongFormError != 1):
                self.checkBtn['takefocus'] = TRUE
                self.checkBtn['state'] = 'active'


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    ui.net_checker()    # 현재 네트워크가 정상적으로 연결되어 있는지 확인 (최초 호출)
    Dialog.show()
    
    sys.exit(app.exec_())
