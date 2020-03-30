import json
from django.shortcuts import render, HttpResponse
from .models import TestColorS, TestColorA, AutomationProgram
import openpyxl
# Create your views here.

def qualityauto(request):
	datasS = TestColorS.objects.filter(is_deleted=True)
	testcolorslist = [data.version for data in datasS]
	paths = ['/home/qceo/QCEO/report/'+str(data.case) for data in datasS]
	testcolors = []

	for x in range(0, len(datasS)):
		ver = str(testcolorslist[x])
		npath = str(datasS[x].case)
		words = []
		exceldoc = openpyxl.load_workbook(paths[x])
		ws = exceldoc.active
		for x in range(2, 1048576):
			if ws['F'+str(x)].value:
				words.append(ws['F'+str(x)].value)
			else:
				break
		passcnt, failcnt = 0, 0

		for y in words:
			if y == 'PASS':
				passcnt+=1
			else:
				failcnt+=1
		total = passcnt+failcnt
		ret = [ver, passcnt, failcnt, total, npath]
		testcolors.append(ret)

	

	datasA = TestColorA.objects.filter(is_deleted=True)

	testcoloralist = [data.version for data in datasA]
	paths = ['/home/qceo/QCEO/report/'+str(data.case) for data in datasA]

	testcolora = []

	for x in range(0, len(datasA)):
		ver = str(testcoloralist[x])
		npath = str(datasA[x].case)
		words = []
		exceldoc = openpyxl.load_workbook(paths[x])
		ws = exceldoc.active
		for x in range(2, 1048576):
			if ws['F'+str(x)].value:
				words.append(ws['F'+str(x)].value)
			else:
				break
		passcnt, failcnt = 0, 0

		for y in words:
			if y == 'PASS':
				passcnt+=1
			else:
				failcnt+=1
		total = passcnt+failcnt
		ret = [ver, passcnt, failcnt, total, npath]
		testcolora.append(ret)

	aidx=str(100)


	programset = AutomationProgram.objects.all()

	return render(request, 'qualityauto/qualityauto.html', {'testcolors':testcolors, 'testcolora':testcolora, 'aidx':aidx, 'programset':programset	})